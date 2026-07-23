from __future__ import annotations

from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from dashboard.data import assess_data_quality, load_books

HEADER_FILL = PatternFill("solid", fgColor="1E3A8A")
HEADER_FONT = Font(color="FFFFFF", bold=True)


def protect_spreadsheet_text(value: object) -> object:
    """Prevent externally collected text from becoming an Excel formula."""
    if not isinstance(value, str):
        return value
    if value.lstrip().startswith(("=", "+", "-", "@", "\t", "\r")):
        return f"'{value}"
    return value


def protect_tabular_text(frame: pd.DataFrame) -> pd.DataFrame:
    """Return a copy with formula-like text neutralized for Excel or CSV."""
    safe_frame = frame.copy()
    for column in safe_frame.select_dtypes(include=["object", "string"]).columns:
        safe_frame[column] = safe_frame[column].map(protect_spreadsheet_text)
    return safe_frame


def export_excel(
    database_path: str | Path,
    output_path: str | Path,
) -> Path:
    books = load_books(database_path)
    if books.empty:
        raise ValueError("No book records were found in the SQLite database")

    safe_books = protect_tabular_text(books)

    output = Path(output_path)
    output.parent.mkdir(parents=True, exist_ok=True)

    summary = pd.DataFrame(
        [
            ("Products", len(books)),
            ("Categories", books["category"].nunique()),
            ("Average price", round(books["price_gbp"].mean(), 2)),
            ("Median price", round(books["price_gbp"].median(), 2)),
            ("Average rating", round(books["rating"].mean(), 2)),
            ("In-stock products", int((books["stock_count"] > 0).sum())),
        ],
        columns=["Metric", "Value"],
    )
    category_summary = (
        books.groupby("category", as_index=False)
        .agg(
            products=("source_id", "count"),
            average_price=("price_gbp", "mean"),
            average_rating=("rating", "mean"),
        )
        .sort_values(["products", "category"], ascending=[False, True])
    )
    quality_summary = assess_data_quality(books)

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        safe_books.to_excel(writer, sheet_name="Products", index=False)
        summary.to_excel(writer, sheet_name="Summary", index=False)
        category_summary.to_excel(
            writer,
            sheet_name="Category Summary",
            index=False,
        )
        quality_summary.to_excel(
            writer,
            sheet_name="Data Quality",
            index=False,
        )

        for worksheet in writer.book.worksheets:
            worksheet.freeze_panes = "A2"
            worksheet.auto_filter.ref = worksheet.dimensions
            for cell in worksheet[1]:
                cell.fill = HEADER_FILL
                cell.font = HEADER_FONT
                cell.alignment = Alignment(horizontal="center")

            for column_cells in worksheet.columns:
                values = [str(cell.value or "") for cell in column_cells]
                width = min(max(len(value) for value in values) + 2, 48)
                worksheet.column_dimensions[
                    get_column_letter(column_cells[0].column)
                ].width = max(width, 12)

        products_sheet = writer.book["Products"]
        price_headers = {
            "price_gbp",
            "price_excl_tax_gbp",
            "price_incl_tax_gbp",
            "tax_gbp",
        }
        for cell in products_sheet[1]:
            if cell.value in price_headers:
                for data_cell in products_sheet.iter_cols(
                    min_col=cell.column,
                    max_col=cell.column,
                    min_row=2,
                ):
                    for price_cell in data_cell:
                        price_cell.number_format = '£0.00'

        summary_sheet = writer.book["Summary"]
        for row in range(2, summary_sheet.max_row + 1):
            if summary_sheet.cell(row=row, column=1).value in {
                "Average price",
                "Median price",
            }:
                summary_sheet.cell(row=row, column=2).number_format = "£0.00"

        category_sheet = writer.book["Category Summary"]
        category_headers = {
            cell.value: cell.column for cell in category_sheet[1]
        }
        for row in range(2, category_sheet.max_row + 1):
            category_sheet.cell(
                row=row,
                column=category_headers["average_price"],
            ).number_format = "£0.00"
            category_sheet.cell(
                row=row,
                column=category_headers["average_rating"],
            ).number_format = "0.00"

    return output
