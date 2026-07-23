
import pandas as pd
from openpyxl import load_workbook

from catalog_scraper.pipelines import NormalizeBookPipeline, SQLiteBookPipeline
from dashboard.data import assess_data_quality, filter_books, load_books
from exporter.excel import export_excel
from tests.test_pipelines import raw_book


def database_with_books(tmp_path):
    database = tmp_path / "books.sqlite"
    pipeline = SQLiteBookPipeline(str(database))
    pipeline.open_spider()
    first = raw_book("book-1")
    first.update({"title": "Python Basics", "category": "Technology", "rating": 5})
    second = raw_book("book-2")
    second.update({"title": "Mountain Travel", "category": "Travel", "rating": 3})
    normalizer = NormalizeBookPipeline()
    pipeline.process_item(normalizer.process_item(first))
    pipeline.process_item(normalizer.process_item(second))
    pipeline.close_spider()
    return database


def test_load_and_filter_books(tmp_path):
    database = database_with_books(tmp_path)
    books = load_books(database)

    result = filter_books(
        books,
        search="python",
        categories=["Technology"],
        minimum_rating=4,
        price_range=(0, 20),
    )
    assert result["source_id"].tolist() == ["book-1"]


def test_filter_books_does_not_mutate_input():
    books = pd.DataFrame(
        [
            {
                "title": "Example",
                "description": "",
                "category": "Test",
                "rating": 4,
                "price_gbp": 10.0,
            }
        ]
    )
    original = books.copy(deep=True)
    filter_books(books, search="example")
    pd.testing.assert_frame_equal(books, original)


def test_export_excel_creates_expected_sheets(tmp_path):
    database = database_with_books(tmp_path)
    output = export_excel(database, tmp_path / "report.xlsx")

    with pd.ExcelFile(output) as workbook:
        assert workbook.sheet_names == [
            "Products",
            "Summary",
            "Category Summary",
            "Data Quality",
        ]


def test_empty_database_path_returns_empty_dataframe(tmp_path):
    assert load_books(tmp_path / "missing.sqlite").empty


def test_data_quality_checks_pass_for_valid_records(tmp_path):
    books = load_books(database_with_books(tmp_path))
    quality = assess_data_quality(books)

    assert quality["Status"].tolist() == ["Passed"] * 5
    assert quality["Issues"].sum() == 0


def test_data_quality_detects_duplicate_ids_and_invalid_values():
    books = pd.DataFrame(
        [
            {
                "source_id": "same",
                "title": "Valid",
                "category": "Test",
                "price_gbp": 10,
                "rating": 5,
                "product_url": "https://example.test/same",
                "scraped_at": "2026-07-23T00:00:00+00:00",
            },
            {
                "source_id": "same",
                "title": "",
                "category": "Test",
                "price_gbp": -1,
                "rating": 7,
                "product_url": "https://example.test/same",
                "scraped_at": "2026-07-23T00:00:00+00:00",
            },
        ]
    )
    quality = assess_data_quality(books).set_index("Check")

    assert quality.loc["Required fields", "Issues"] == 1
    assert quality.loc["Unique product IDs", "Issues"] == 1
    assert quality.loc["Unique product URLs", "Issues"] == 1
    assert quality.loc["Valid prices", "Issues"] == 1
    assert quality.loc["Ratings between 1 and 5", "Issues"] == 1


def test_excel_export_neutralizes_formula_like_text(tmp_path):
    database = database_with_books(tmp_path)
    pipeline = SQLiteBookPipeline(str(database))
    pipeline.open_spider()
    malicious = raw_book("book-3")
    malicious["title"] = "=HYPERLINK(\"https://example.test\")"
    pipeline.process_item(NormalizeBookPipeline().process_item(malicious))
    pipeline.close_spider()

    output = export_excel(database, tmp_path / "safe-report.xlsx")
    workbook = load_workbook(output, data_only=False)
    products = workbook["Products"]
    title_column = next(
        cell.column for cell in products[1] if cell.value == "title"
    )
    values = [
        products.cell(row=row, column=title_column).value
        for row in range(2, products.max_row + 1)
    ]

    assert "'=HYPERLINK" in next(value for value in values if "HYPERLINK" in value)
